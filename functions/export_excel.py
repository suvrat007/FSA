import io
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from functions.assessment import CompanyAssessment, assess_company
from functions.assumptions import ForecastAssumptions
from functions.config import APP_NAME
from functions.datamodel import FinancialDataModel
from functions.narrative import SECTION_ORDER, generate_executive_narrative
from functions.quality import evaluate_red_flags
from functions.ratios import compute_all_ratios
from functions.scenarios import build_scenario_models, compile_scenario_summary_table

HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="0F172A")
LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="000000")
BODY_FONT = Font(name="Calibri", size=10, color="000000")
ROW_BORDER = Border(
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

NUMBER_FORMAT = "#,##0.00"
MIN_COLUMN_WIDTH = 12
MAX_COLUMN_WIDTH = 60
NARRATIVE_WRAP_WIDTH = 110

RED_FLAG_COLUMNS = ["rule_id", "rule_name", "severity", "observation", "finance_reason"]
RED_FLAG_HEADERS = ["Rule", "Finding", "Severity", "Observation", "Accounting rationale"]


def export_model_to_excel(
    model: FinancialDataModel,
    forecast_model: Optional[FinancialDataModel] = None,
    assumptions: Optional[ForecastAssumptions] = None,
) -> bytes:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    assessment = assess_company(model)

    _write_summary_sheet(workbook, model, assessment)
    _write_dataframe(workbook, "Assessment", assessment.summary_frame(), include_index=False)
    _write_dataframe(workbook, "Data Quality", model.provenance_frame(), include_index=False)

    reporting_model = forecast_model if forecast_model is not None else model

    _write_dataframe(workbook, "Income Statement", reporting_model.income_statement)
    _write_dataframe(workbook, "Balance Sheet", reporting_model.balance_sheet)
    _write_dataframe(workbook, "Cash Flow", reporting_model.cash_flow)
    _write_dataframe(workbook, "Ratios", compute_all_ratios(reporting_model))

    red_flags = evaluate_red_flags(model)
    if red_flags:
        flags_df = pd.DataFrame(red_flags)[RED_FLAG_COLUMNS]
        flags_df.columns = RED_FLAG_HEADERS
        _write_dataframe(workbook, "Accounting Quality", flags_df, include_index=False)

    if assumptions is not None:
        scenarios = build_scenario_models(model, assumptions)
        _write_dataframe(
            workbook,
            "Scenarios",
            compile_scenario_summary_table(scenarios),
            include_index=False,
        )

    buffer = io.BytesIO()
    workbook.save(buffer)

    return buffer.getvalue()


def _write_summary_sheet(
    workbook: openpyxl.Workbook,
    model: FinancialDataModel,
    assessment: CompanyAssessment,
) -> None:
    sheet = workbook.create_sheet(title="Summary")

    sheet.cell(row=1, column=1, value=f"{APP_NAME} Investment Research Report").font = TITLE_FONT
    sheet.cell(row=2, column=1, value=f"{model.company_name} ({model.ticker})").font = LABEL_FONT
    sheet.cell(
        row=3,
        column=1,
        value=f"Sector: {model.metadata.get('Sector', 'N/A')}    Currency: {model.currency}"
        f"    Source: {model.source}",
    ).font = BODY_FONT
    sheet.cell(
        row=4,
        column=1,
        value=f"Overall verdict: {assessment.grade} ({assessment.score:.0f} out of 100)",
    ).font = LABEL_FONT

    narratives = generate_executive_narrative(model)
    row = 6

    for section in SECTION_ORDER:
        if section not in narratives:
            continue

        row = _write_paragraph(sheet, row, section, narratives[section])

    row = _write_paragraph(
        sheet, row, "Data quality and assumptions", " ".join(assessment.caveats)
    )

    sheet.column_dimensions["A"].width = NARRATIVE_WRAP_WIDTH


def _write_paragraph(sheet, row: int, heading: str, text: str) -> int:
    sheet.cell(row=row, column=1, value=heading).font = LABEL_FONT
    row += 1

    for line in _wrap(text, NARRATIVE_WRAP_WIDTH):
        sheet.cell(row=row, column=1, value=line).font = BODY_FONT
        row += 1

    return row + 1


def _write_dataframe(
    workbook: openpyxl.Workbook,
    sheet_title: str,
    df: pd.DataFrame,
    include_index: bool = True,
) -> None:
    if df is None or df.empty:
        return

    sheet = workbook.create_sheet(title=sheet_title)
    export_df = df.reset_index() if include_index else df

    for row_index, row in enumerate(dataframe_to_rows(export_df, index=False, header=True), 1):
        sheet.append(row)

        for column_index, cell in enumerate(sheet[row_index], 1):
            if row_index == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="left" if column_index == 1 else "center")
                continue

            cell.font = BODY_FONT
            cell.border = ROW_BORDER

            if isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right")

    _autosize_columns(sheet)


def _autosize_columns(sheet) -> None:
    for column in sheet.columns:
        longest = max(len(str(cell.value or "")) for cell in column)
        width = min(max(longest + 3, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


def _wrap(text: str, width: int) -> list:
    words = text.split()
    lines = []
    current = ""

    for word in words:
        candidate = f"{current} {word}".strip()

        if len(candidate) > width and current:
            lines.append(current)
            current = word
        else:
            current = candidate

    if current:
        lines.append(current)

    return lines
